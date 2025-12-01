# bot.py
import os
import logging
from logging.handlers import RotatingFileHandler
from io import BytesIO
from datetime import datetime
from typing import List, Any
from dotenv import load_dotenv
import pathlib
import uuid

import openpyxl
import xlrd
from rapidfuzz import process, fuzz

from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
    PicklePersistence,
)

# Load local .env for dev if present
load_dotenv()

# -------------------------
# Configuration (env-driven)
# -------------------------
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "uploads")
PERSISTENCE_FILE = os.getenv("PERSISTENCE_FILE", "bot_data.pkl")
UPLOAD_MAX_MB = int(os.getenv("UPLOAD_MAX_MB", "8"))  # max upload size in MB
UPLOAD_MAX_BYTES = UPLOAD_MAX_MB * 1024 * 1024
MAX_ROWS = int(os.getenv("MAX_ROWS", "200000"))  # protect from enormous sheets
LOG_FILE = os.getenv("LOG_FILE", "bot.log")
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET")  # optional secret for webhook path

# Ensure upload dir exists
pathlib.Path(UPLOAD_DIR).mkdir(parents=True, exist_ok=True)

# -------------------------
# Logging (console + rotating file)
# -------------------------
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Console handler (already configured by basicConfig below for compatibility)
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# Rotating file handler
file_handler = RotatingFileHandler(LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=3)
file_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))
logger.addHandler(file_handler)

# Conversation states
WAITING_FILE, WAITING_COLUMN, WAITING_QUERY, WAITING_SELECT = range(4)
WAITING_CREATE_COLUMNS = 100

# -------------------------
# Utility helpers
# -------------------------
def is_integer_like_number(v: Any) -> bool:
    try:
        if isinstance(v, float):
            return v.is_integer()
        if isinstance(v, int):
            return True
    except Exception:
        return False
    return False

def format_number_no_scientific(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        else:
            return format(v, "f").rstrip("0").rstrip(".")
    return str(v)

def parse_possible_date(value: Any):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    try:
        if isinstance(value, (int, float)):
            if 1000 < value < 100000:
                try:
                    dt = xlrd.xldate.xldate_as_datetime(value, 0)
                    return dt
                except Exception:
                    pass
    except Exception:
        pass

    if isinstance(value, str):
        s = value.strip()
        fmts = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%b-%Y",
            "%d %b %Y",
            "%m/%d/%Y",
            "%Y.%m.%d %H:%M:%S",
            "%Y.%m.%d"
        ]
        for f in fmts:
            try:
                return datetime.strptime(s, f)
            except Exception:
                continue

        try:
            if "t" in s.lower():
                return datetime.fromisoformat(s)
        except Exception:
            pass

    return None

def format_cell_for_output(value: Any):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, (int, float)) and is_integer_like_number(value):
        return str(int(value))
    possible_date = parse_possible_date(value)
    if possible_date:
        return possible_date.strftime("%d-%m-%Y")
    if isinstance(value, str):
        return value.strip()
    return str(value)

def load_excel_clean_from_bytes(file_bytes: bytes, file_name: str) -> openpyxl.Workbook:
    """
    Load workbook from bytes and return a clean openpyxl Workbook containing values-only.
    """
    lower = file_name.lower()
    if lower.endswith(".xlsx"):
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        clean_wb = openpyxl.Workbook()
        clean_ws = clean_wb.active
        for row in ws.iter_rows(values_only=True):
            clean_ws.append([cell for cell in row])
        return clean_wb

    if lower.endswith(".xls"):
        try:
            book = xlrd.open_workbook(file_contents=file_bytes)
            sheet = book.sheet_by_index(0)
            clean_wb = openpyxl.Workbook()
            clean_ws = clean_wb.active
            for r in range(sheet.nrows):
                clean_ws.append(sheet.row_values(r))
            return clean_wb

        except xlrd.biffh.XLRDError as e:
            if "xlsx file; not supported" in str(e).lower():
                wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                ws = wb.active
                clean_wb = openpyxl.Workbook()
                clean_ws = clean_wb.active
                for row in ws.iter_rows(values_only=True):
                    clean_ws.append([cell for cell in row])
                return clean_wb
            raise

    raise ValueError("Unsupported file type. Send .xls or .xlsx")

def save_uploaded_file(bytes_data: bytes, original_name: str) -> str:
    """
    Save uploaded file bytes to UPLOAD_DIR and return path.
    Filename uses uuid to avoid collisions.
    """
    uid = uuid.uuid4().hex
    safe_name = "".join(c for c in original_name if c.isalnum() or c in "._- ")[:100]
    filename = f"{uid}_{safe_name}"
    path = os.path.join(UPLOAD_DIR, filename)
    with open(path, "wb") as f:
        f.write(bytes_data)
    return path

# -------------------------
# Handlers (preserve your original flows)
# -------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Send an Excel file (.xls or .xlsx). I'll strip formatting and work with values-only.\n\n"
        "Flow:\n"
        "1) You upload file\n"
        "2) You pick a column number\n"
        "3) You type a partial search string\n"
        "4) I show matching values\n"
        "5) I return filtered rows\n\n"
        "Send /cancel to stop."
    )
    return WAITING_FILE

async def receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        doc = update.message.document
        if not doc:
            await update.message.reply_text("No file detected. Send .xls or .xlsx.")
            return WAITING_FILE

        # File size guard (Telegram Document has file_size attribute)
        file_size = getattr(doc, "file_size", 0) or 0
        if file_size and file_size > UPLOAD_MAX_BYTES:
            await update.message.reply_text(
                f"File too large ({file_size/1024/1024:.1f} MB). "
                f"Max allowed is {UPLOAD_MAX_MB} MB."
            )
            return WAITING_FILE

        fname = doc.file_name
        if not (fname.lower().endswith(".xls") or fname.lower().endswith(".xlsx")):
            await update.message.reply_text("Please send only .xls or .xlsx files.")
            return WAITING_FILE

        tgfile = await context.bot.get_file(doc.file_id)
        fb = await tgfile.download_as_bytearray()

        # Save raw uploaded bytes to persistent upload dir and record path in user_data
        saved_path = save_uploaded_file(bytes(fb), fname)

        # Load workbook values-only to inspect columns/rows and enforce row limits
        wb = load_excel_clean_from_bytes(bytes(fb), fname)
        ws = wb.active

        if ws.max_row - 1 > MAX_ROWS:  # subtract header row
            await update.message.reply_text(
                f"Workbook has {ws.max_row - 1} data rows, exceeds limit of {MAX_ROWS}. "
                "Please upload a smaller file."
            )
            # Optionally remove saved file to avoid storing huge uploads
            try:
                os.remove(saved_path)
            except Exception:
                logger.exception("Failed to remove oversized uploaded file: %s", saved_path)
            return WAITING_FILE

        # Persist only lightweight info (do NOT put workbook objects in user_data)
        context.user_data["file_path"] = saved_path
        context.user_data["file_name"] = fname
        context.user_data["max_col"] = ws.max_column

        headers = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(1, col).value
            if v is None:
                headers.append(f"{col}. Column {openpyxl.utils.get_column_letter(col)}")
            else:
                headers.append(f"{col}. {str(v)}")

        await update.message.reply_text(
            "File loaded. Columns:\n\n" + "\n".join(headers) +
            f"\n\nReply with column number (1 - {ws.max_column})."
        )
        return WAITING_COLUMN

    except Exception as e:
        logger.exception("Error in receive_file")
        await update.message.reply_text("Error while processing file. Make sure it's a valid .xls or .xlsx.")
        return WAITING_FILE

async def receive_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    try:
        c = int(txt)
    except Exception:
        await update.message.reply_text("Send a valid column number.")
        return WAITING_COLUMN

    if "max_col" not in context.user_data:
        await update.message.reply_text("Session expired or no file loaded. Send a file first (/start).")
        return ConversationHandler.END

    if c < 1 or c > context.user_data["max_col"]:
        await update.message.reply_text("Invalid column number.")
        return WAITING_COLUMN

    context.user_data["col"] = c
    await update.message.reply_text("Enter a search string.")
    return WAITING_QUERY

async def receive_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.message.text.strip()
    if not query:
        await update.message.reply_text("Search string cannot be empty.")
        return WAITING_QUERY

    file_path = context.user_data.get("file_path")
    if not file_path or not os.path.exists(file_path):
        await update.message.reply_text("Uploaded file not found (session expired). Please upload again.")
        return ConversationHandler.END

    # Re-load workbook from saved file (safe; keeps user_data pickleable)
    with open(file_path, "rb") as f:
        wb = load_excel_clean_from_bytes(f.read(), context.user_data.get("file_name", "file.xlsx"))
    ws = wb.active
    col = context.user_data.get("col")
    if not col:
        await update.message.reply_text("Column not selected. Start again (/start).")
        return ConversationHandler.END

    seen = {}
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, col).value
        if raw is None:
            continue
        norm = str(raw).strip().lower()
        if query.lower() in norm:
            if norm not in seen:
                seen[norm] = str(raw).strip()

    if not seen:
        await update.message.reply_text("No matches found.")
        return ConversationHandler.END

    choices = list(seen.values())
    ranked = process.extract(query, choices, scorer=fuzz.partial_ratio, limit=200)
    ordered = [t[0] for t in ranked]

    presented = ordered[:40]
    context.user_data["candidates"] = presented

    lines = [f"Found {len(seen)} values. Showing top {len(presented)}:"]
    for i, v in enumerate(presented, 1):
        lines.append(f"{i}. {v}")
    lines.append("\nReply with number (0 to cancel).")

    await update.message.reply_text("\n".join(lines))
    return WAITING_SELECT

async def receive_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    if not txt.isdigit():
        await update.message.reply_text("Enter a valid number.")
        return WAITING_SELECT

    idx = int(txt)
    if idx == 0:
        await update.message.reply_text("Cancelled.")
        return ConversationHandler.END

    candidates: List[str] = context.user_data.get("candidates", [])
    if idx < 1 or idx > len(candidates):
        await update.message.reply_text("Invalid selection.")
        return WAITING_SELECT

    chosen = candidates[idx - 1]
    context.user_data["chosen_value"] = chosen

    file_path = context.user_data.get("file_path")
    if not file_path or not os.path.exists(file_path):
        await update.message.reply_text("Uploaded file not found (session expired). Please upload again.")
        return ConversationHandler.END

    with open(file_path, "rb") as f:
        wb = load_excel_clean_from_bytes(f.read(), context.user_data.get("file_name", "file.xlsx"))
    ws = wb.active
    col = context.user_data.get("col")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Filtered"

    header_row = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        header_row.append(str(v) if v else "")
    out_ws.append(header_row)

    match_count = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, col).value
        if val is None:
            continue
        if chosen.lower() in str(val).strip().lower():
            formatted = [format_cell_for_output(ws.cell(r, c).value)
                         for c in range(1, ws.max_column + 1)]
            out_ws.append(formatted)
            match_count += 1

    if match_count == 0:
        await update.message.reply_text("Unexpected: no rows matched.")
        return ConversationHandler.END

    for c in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(c)
        max_len = 0
        for row in out_ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        out_ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)

    output = BytesIO()
    out_wb.save(output)
    output.seek(0)

    base = os.path.splitext(context.user_data.get("file_name", "filtered"))[0]
    # make filename safe
    chosen_safe = "".join(ch for ch in chosen[:30] if ch.isalnum() or ch in "._- ").replace(" ", "_")
    out_name = f"{base}_filtered_by_{col}_{chosen_safe}.xlsx"

    await update.message.reply_document(
        document=output,
        filename=out_name,
        caption=f"Filtered file ready. {match_count} rows matched."
    )
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Operation cancelled.")
    return ConversationHandler.END

# Create-template handlers
async def create_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Send column names separated by commas.\nExample: name, phone, address"
    )
    return WAITING_CREATE_COLUMNS

async def receive_create_columns(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("Column list cannot be empty.")
        return WAITING_CREATE_COLUMNS

    cols = [c.strip() for c in text.split(",") if c.strip()]
    if not cols:
        await update.message.reply_text("Invalid column list.")
        return WAITING_CREATE_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(cols)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"empty_template_{timestamp}.xlsx"

    await update.message.reply_document(
        document=output,
        filename=filename,
        caption="Empty Excel template created."
    )

    return ConversationHandler.END

# -------------------------
# Global error handler
# -------------------------
async def global_error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    # Log exception with stacktrace
    logger.exception("Unhandled exception: %s", context.error)
    # Try to tell the user something friendly
    try:
        if isinstance(update, Update) and update.effective_chat:
            await context.bot.send_message(update.effective_chat.id, "Internal error occurred. Try again later.")
    except Exception:
        logger.exception("Failed to notify user about exception")

# -------------------------
# Main application: build handlers & start either polling or webhook
# -------------------------
def build_application(token: str) -> Application:
    # Persistence for user_data, chat_data, bot_data
    persistence = PicklePersistence(filepath=PERSISTENCE_FILE)
    app = Application.builder().token(token).persistence(persistence).build()

    # Filtering workflow (original)
    filter_conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Document.ALL, receive_file)
        ],
        states={
            WAITING_FILE: [MessageHandler(filters.Document.ALL, receive_file)],
            WAITING_COLUMN: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_column)],
            WAITING_QUERY: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_query)],
            WAITING_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_select)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True
    )

    create_conv = ConversationHandler(
        entry_points=[CommandHandler("create", create_start)],
        states={
            WAITING_CREATE_COLUMNS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_create_columns)
            ]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True
    )

    app.add_handler(filter_conv)
    app.add_handler(create_conv)

    # Health handler
    async def health(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text("OK")

    app.add_handler(CommandHandler("health", health))

    # Global error handler
    app.add_error_handler(global_error_handler)

    return app

def main():
    TOKEN = os.getenv("TOKEN")
    if not TOKEN:
        raise SystemExit("TOKEN not set in environment.")

    app = build_application(TOKEN)

    # Webhook configuration:
    WEBHOOK_BASE = os.getenv("WEBHOOK_BASE_URL")  # e.g. "https://your-render-service.onrender.com"
    PORT = int(os.getenv("PORT", "8443"))

    if WEBHOOK_BASE:
        # Clean base (no trailing slash)
        WEBHOOK_BASE = WEBHOOK_BASE.rstrip("/")
        # Use webhook secret if provided, otherwise use token-derived path (less ideal)
        webhook_token_path = WEBHOOK_SECRET if WEBHOOK_SECRET else TOKEN
        webhook_path = f"/webhook/{webhook_token_path}"
        webhook_url = f"{WEBHOOK_BASE}{webhook_path}"
        logger.info("Starting in WEBHOOK mode. Webhook URL: %s", webhook_url)

        try:
            # run_webhook will start aiohttp server and call setWebhook for Telegram
            app.run_webhook(
                listen="0.0.0.0",
                port=PORT,
                url_path=webhook_path.lstrip("/"),
                webhook_url=webhook_url
            )
        except Exception as e:
            logger.exception("Failed to start webhook mode: %s", e)
            raise
    else:
        # Polling fallback for local development / simple always-on server
        logger.info("No WEBHOOK_BASE_URL configured. Starting in polling mode.")
        app.run_polling()

if __name__ == "__main__":
    main()
